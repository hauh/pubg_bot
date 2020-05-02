'''Reading and creating excel files'''

from io import BytesIO, BufferedWriter
from zipfile import BadZipFile

import openpyxl
from openpyxl.styles.numbers import FORMAT_NUMBER

##############################

PLAYERS_TABLE_COLUMNS = [  # with width
	('PUBG ID', 12),
	('PUBG Username', 20),
	('Telegram ID', 12),
	('Telegram Username', 20),
	('Place', 7),
	('Kills', 7),
]


def create_table(players):
	players_table = openpyxl.Workbook()
	sheet = players_table.active

	for index, column_name in enumerate(PLAYERS_TABLE_COLUMNS, start=1):
		name, width = column_name
		sheet.cell(1, index, name)
		sheet.column_dimensions[str(chr(64 + index))].width = width

	for row, player in enumerate(players, start=2):
		for column, player_data in enumerate(player.values(), start=1):
			sheet.cell(row, column, player_data).number_format = FORMAT_NUMBER

	virtual_file = BytesIO()
	players_table.save(virtual_file)
	virtual_file.seek(0)
	return virtual_file


def read_table(winners_file):
	winners_bytes = BytesIO()
	writer = BufferedWriter(
		winners_bytes, buffer_size=int(winners_file.file_size))
	winners_file.download(out=writer)
	writer.flush()
	winners_bytes.seek(0)
	try:
		return openpyxl.load_workbook(filename=winners_bytes).active
	except BadZipFile:
		return None


def get_winners(results):
	for row in results.iter_rows(min_row=2):
		if place := row[4].value:
			yield row[2].row, row[2].value, place


def get_killers(results):
	for row in results.iter_rows(min_row=2):
		if kills := row[5].value:
			yield row[2].row, row[2].value, kills
